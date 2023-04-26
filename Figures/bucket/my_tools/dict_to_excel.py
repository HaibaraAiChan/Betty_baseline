import openpyxl

# Your dictionary
# data = {
#     "Name": ["Alice", "Bob", "Charlie"],
#     "Age": [25, 30, 22],
#     "City": ["New York", "San Francisco", "Los Angeles"]
# }
data = {1: 16700, 2: 16060, 3: 13581, 4: 11382, 5: 9803, 6: 8141, 7: 7143, 8: 6274, 9: 5511, 10: 60057}
data = {1: 13428, 2: 11706, 3: 9277, 4: 7320, 5: 6222, 6: 4868, 7: 4045, 8: 3472, 9: 2976, 10: 2599, 11: 2203, 12: 1937, 13: 1656, 14: 1434, 15: 1289, 16: 1111, 17: 1030, 18: 948, 19: 836, 20: 807, 21: 717, 22: 556, 23: 519, 24: 525, 25: 9460}
data = {1: 15521, 2: 14725, 3: 12212, 4: 10016, 5: 8580, 6: 6908, 7: 5873, 8: 5129, 9: 4387, 10: 36078}
data = {1: 13428, 2: 11706, 3: 9277, 4: 7320, 5: 6222, 6: 4868, 7: 4045, 8: 3472, 9: 2976, 10: 2599, 11: 2203, 12: 1937}
data = {1: 3930, 2: 5869, 3: 6607, 4: 6704, 5: 6629, 6: 6136, 7: 5750, 8: 5307, 9: 4820, 10: 58453}
data = {13: 1656, 14: 1434, 15: 1289, 16: 1111, 17: 1030, 18: 948, 19: 836, 20: 807, 21: 717, 22: 556, 23: 519, 24: 525, 25: 9460}

data = {1: 15521, 2: 14725, 3: 12212, 4: 10016, 5: 8580, 6: 6908, 7: 5873, 8: 5129, 9: 4387, 10: 36078}
data = {1: 13428, 2: 11706, 3: 9277, 4: 7320, 5: 6222, 6: 4868, 7: 4045, 8: 3472, 9: 2976, 10: 2599, 11: 2203, 12: 1937}


data = {1: 2008, 2: 3194, 3: 3811, 4: 4194, 5: 4275, 6: 4196, 7: 4026, 8: 3932, 9: 3647, 10: 50625}
data = {13: 812, 14: 742, 15: 648, 16: 561, 17: 515, 18: 470, 19: 434, 20: 383, 21: 378, 22: 257, 23: 284, 24: 266, 25: 4694}


data = {1: 1922, 2: 3118, 3: 3788, 4: 4080, 5: 4293, 6: 4180, 7: 4081, 8: 3896, 9: 3669, 10: 50598}
data = {13: 844, 14: 692, 15: 641, 16: 550, 17: 515, 18: 478, 19: 402, 20: 424, 21: 339, 22: 299, 23: 235, 24: 259, 25: 4766}

# Create a new Excel workbook and add a worksheet
wb = openpyxl.Workbook()
ws = wb.active
# Write the header (keys) to the first colomn
for idx, key in enumerate(data.keys(), start=1):
    ws.cell(row=idx, column=2, value=key)
# Write the data (values) to the following rows
for row_idx, row_data in enumerate(zip(data.values()), start=1):
    for col_idx, cell_data in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=cell_data)
# Save the workbook to a file
# wb.save("nb_3_arxiv_block_01.xlsx")
# wb.save("nb_3_arxiv_block_11.xlsx")
wb.save("arxiv_block_21.xlsx")
print("Dictionary has been successfully transferred to my_data.xlsx")
